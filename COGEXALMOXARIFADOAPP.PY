# app.py

import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# -------------------- CONFIGURAÇÕES INICIAIS --------------------
st.set_page_config(page_title="COGEX Almoxarifado", layout="wide")
st.title("📦 COGEX ALMOXARIFADO")
st.markdown("**Sistema Integrado Google Sheets - Pedido de Material com Imagens, Filtros e Preditivos**")

# -------------------- CARREGAMENTO DE DADOS --------------------
@st.cache_data(show_spinner="Carregando dados do Google Sheets...")
def load_data():
    url_inventory = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSeWsxmLFzuWsa2oggpQb6p5SFapxXHcWaIl0Jjf2wAezvMgAV9XCc1r7fSSzRWTCgjk9eqREgWlrzp/pub?gid=1710164548&single=true&output=csv'
    url_items = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSeWsxmLFzuWsa2oggpQb6p5SFapxXHcWaIl0Jjf2wAezvMgAV9XCc1r7fSSzRWTCgjk9eqREgWlrzp/pub?gid=1011017078&single=true&output=csv'

    inventory = pd.read_csv(url_inventory)
    inventory['DateTime'] = pd.to_datetime(inventory['DateTime'], errors='coerce')
    items = pd.read_csv(url_items)

    items['Image_Link'] = items['Image'].apply(lambda x: f'https://drive.google.com/uc?export=view&id=17DLW40Xz_UiOS-YVR-BzlYffV6mQsQvp/{x}')

    return items, inventory

items_df, inventory_df = load_data()

# -------------------- PREPARAÇÃO DOS DADOS --------------------
merged_df = pd.merge(inventory_df, items_df, on='Item ID', how='left')
merged_df['Ano'] = merged_df['DateTime'].dt.year
merged_df['Mês'] = merged_df['DateTime'].dt.month
merged_df['Semana'] = merged_df['DateTime'].dt.isocalendar().week

# -------------------- CONSUMO MÉDIO --------------------
def consumo_medio(df, dias):
    data_limite = datetime.now() - timedelta(days=dias)
    consumo = df[(df['DateTime'] >= data_limite) & (df['Amount'] < 0)]
    consumo_agrupado = consumo.groupby(['Item ID', 'Name', 'Image_Link'])['Amount'].sum().abs().reset_index()
    consumo_agrupado.rename(columns={'Amount': f'Consumo Médio {dias} dias'}, inplace=True)
    return consumo_agrupado

consumo_7 = consumo_medio(merged_df, 7)
consumo_15 = consumo_medio(merged_df, 15)
consumo_30 = consumo_medio(merged_df, 30)
consumo_45 = consumo_medio(merged_df, 45)

consumo_total = consumo_7.merge(consumo_15, on=['Item ID', 'Name', 'Image_Link'], how='outer')\
                        .merge(consumo_30, on=['Item ID', 'Name', 'Image_Link'], how='outer')\
                        .merge(consumo_45, on=['Item ID', 'Name', 'Image_Link'], how='outer').fillna(0)

estoque_atual = inventory_df.groupby('Item ID')['Amount'].sum().reset_index()
estoque_atual = pd.merge(estoque_atual, items_df[['Item ID', 'Name']], on='Item ID', how='left')

pedido_material = pd.merge(consumo_total, estoque_atual, on=['Item ID', 'Name'], how='left')
pedido_material['Estoque Atual'] = pedido_material['Amount']
pedido_material.drop(columns=['Amount'], inplace=True)

pedido_material['Recomendação Pedido'] = np.where(
    pedido_material['Estoque Atual'] < pedido_material['Consumo Médio 15 dias'],
    'Pedido Necessário',
    'OK'
)

# -------------------- TABS --------------------
tabs = st.tabs(["📋 Tabela & Filtros", "🖼️ Detalhes por Produto", "📊 Estatísticas & Alertas", "📥 Pedido Automático COGEX"])

with tabs[3]:
    st.header("📥 Pedido Automático Almoxarifado COGEX")

    pedido_auto = pedido_material.copy()
    pedido_auto['Pedido 7 dias'] = pedido_auto['Consumo Médio 7 dias'] - pedido_auto['Estoque Atual']
    pedido_auto['Pedido 15 dias'] = pedido_auto['Consumo Médio 15 dias'] - pedido_auto['Estoque Atual']
    pedido_auto['Pedido 30 dias'] = pedido_auto['Consumo Médio 30 dias'] - pedido_auto['Estoque Atual']
    pedido_auto['Pedido 45 dias'] = pedido_auto['Consumo Médio 45 dias'] - pedido_auto['Estoque Atual']

    pedido_auto[pedido_auto.columns[-4:]] = pedido_auto[pedido_auto.columns[-4:]].applymap(lambda x: max(x,0))

    st.dataframe(pedido_auto[['Item ID', 'Name', 'Estoque Atual', 'Pedido 7 dias', 'Pedido 15 dias', 'Pedido 30 dias', 'Pedido 45 dias']])

    st.subheader("📄 Geração de Pedido para Exportação")
    dias_opcao = st.radio("Selecione o Período do Pedido:", options=['7 dias', '15 dias', '30 dias', '45 dias'])

    coluna_pedido = f'Pedido {dias_opcao}'
    pedido_exportar = pedido_auto[['Item ID', 'Name', 'Estoque Atual', coluna_pedido]].rename(columns={coluna_pedido: 'Quantidade Pedido'})

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido_COGEX"

    vin_color = "7B1F2F"
    yellow_fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:D1')
    ws['A1'] = "CORREGEDORIA DO FORO EXTRAJUDICIAL"
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws['A1'].fill = PatternFill(start_color=vin_color, end_color=vin_color, fill_type="solid")

    ws.merge_cells('A2:D2')
    ws['A2'] = "PEDIDO DE MATERIAL AUTOMÁTICO COGEX-MA"
    ws['A2'].font = Font(bold=True, color="000000", size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].fill = yellow_fill

    ws.append([])
    ws.append(["Item ID", "Name", "Estoque Atual", "Quantidade Pedido"])

    for cell in ws[4]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row in pedido_exportar.itertuples(index=False):
        ws.append(row)

    for row in ws.iter_rows(min_row=5, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    footer_row = ws.max_row + 2
    ws.merge_cells(f'A{footer_row}:D{footer_row}')
    ws[f'A{footer_row}'] = "Corregedoria Geral do Foro Extrajudicial - Rua Cumã, nº 300, 1º andar, Edifício Manhattan Center III, Jardim Renascença 2 - São Luís - Maranhão CEP 65.075-700"
    ws[f'A{footer_row}'].alignment = Alignment(horizontal="center")
    ws[f'A{footer_row}'].font = Font(size=9, italic=True)

    excel_output = BytesIO()
    wb.save(excel_output)

    st.download_button(
        label="📥 Baixar Pedido Exportado (XLS)",
        data=excel_output.getvalue(),
        file_name=f'pedido_material_{dias_opcao.replace(" ", "_").lower()}_cogex.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
